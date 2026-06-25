"""
APILedger - 测试数据生成脚本

生成多个 xlsx 文件用于测试导入、冲突检测、extra 列等功能。
运行方式: python test/generate_test_data.py
"""

import csv
import os
from openpyxl import Workbook

TEST_DIR = os.path.dirname(os.path.abspath(__file__))

# ═══════════════════════════════════════════════════════════════
# 共享测试数据
# ═══════════════════════════════════════════════════════════════

# 3天: 2024-06-01 ~ 2024-06-03
DATES = ["2024-06-01", "2024-06-02", "2024-06-03"]

PLATFORMS = ["OpenAI", "Azure", "百度文心"]
PROJECTS = ["智能客服", "代码助手", "文档分析"]
MODELS = ["GPT-4o", "GPT-4o-mini", "GPT-4-Turbo", "文心一言4.0"]
TYPES = ["输入", "输出"]


def make_basic_records():
    """生成基础记录列表，供多个文件复用"""
    records = []
    for date in DATES:
        for plat in PLATFORMS:
            for proj in PROJECTS[:2]:  # 每平台2个项目
                for model in MODELS[:2]:  # 每项目2个模型
                    for typ in TYPES:
                        records.append({
                            "日期": date,
                            "平台": plat,
                            "项目": proj,
                            "模型": model,
                            "类型": typ,
                            "tokens": 5000 + len(records) * 1000,
                            "调用量": 10 + len(records),
                            "金额": round(0.05 + len(records) * 0.01, 2),
                        })
    return records


def write_xlsx(filename, headers, rows):
    """写入 xlsx 文件"""
    wb = Workbook()
    ws = wb.active

    # 表头
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)

    # 数据行
    for row_idx, row in enumerate(rows, 2):
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(h, ""))

    path = os.path.join(TEST_DIR, filename)
    wb.save(path)
    print(f"  已生成: {filename}  ({len(rows)} 行)")


def write_csv(filename, headers, rows, encoding="utf-8-sig"):
    """写入 csv 文件 (默认 utf-8-sig, 兼容 Excel 直接打开)"""
    path = os.path.join(TEST_DIR, filename)
    with open(path, "w", newline="", encoding=encoding) as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
    print(f"  已生成: {filename}  ({len(rows)} 行, {encoding})")


# ═══════════════════════════════════════════════════════════════
# 1. 中文列名 — 常规场景
# ═══════════════════════════════════════════════════════════════

def generate_basic_chinese():
    headers = ["日期", "平台", "项目", "模型", "类型", "tokens", "调用量", "金额"]
    records = [
        # 2024-06-01
        {"日期": "2024-06-01", "平台": "OpenAI", "项目": "智能客服", "模型": "GPT-4o", "类型": "输入", "tokens": 150000, "调用量": 120, "金额": 4.50},
        {"日期": "2024-06-01", "平台": "OpenAI", "项目": "智能客服", "模型": "GPT-4o", "类型": "输出", "tokens": 80000, "调用量": 120, "金额": 3.20},
        {"日期": "2024-06-01", "平台": "Azure", "项目": "代码助手", "模型": "GPT-4-Turbo", "类型": "输入", "tokens": 200000, "调用量": 200, "金额": 2.00},
        {"日期": "2024-06-01", "平台": "Azure", "项目": "代码助手", "模型": "GPT-4-Turbo", "类型": "输出", "tokens": 95000, "调用量": 200, "金额": 2.85},
        # 2024-06-02
        {"日期": "2024-06-02", "平台": "OpenAI", "项目": "智能客服", "模型": "GPT-4o", "类型": "输入", "tokens": 160000, "调用量": 130, "金额": 4.80},
        {"日期": "2024-06-02", "平台": "OpenAI", "项目": "智能客服", "模型": "GPT-4o", "类型": "输出", "tokens": 90000, "调用量": 130, "金额": 3.60},
        {"日期": "2024-06-02", "平台": "百度文心", "项目": "文档分析", "模型": "文心一言4.0", "类型": "输入", "tokens": 300000, "调用量": 300, "金额": 6.00},
        {"日期": "2024-06-02", "平台": "百度文心", "项目": "文档分析", "模型": "文心一言4.0", "类型": "输出", "tokens": 180000, "调用量": 300, "金额": 7.20},
    ]
    write_xlsx("test_basic_chinese.xlsx", headers, records)


# ═══════════════════════════════════════════════════════════════
# 2. 英文列名
# ═══════════════════════════════════════════════════════════════

def generate_basic_english():
    headers = ["Date", "Platform", "Project", "Model", "Type", "Tokens", "Calls", "Cost"]
    records = [
        {"Date": "2024-06-02", "Platform": "OpenAI", "Project": "智能客服", "Model": "GPT-4o-mini", "Type": "输入", "Tokens": 50000, "Calls": 80, "Cost": 0.75},
        {"Date": "2024-06-02", "Platform": "OpenAI", "Project": "智能客服", "Model": "GPT-4o-mini", "Type": "输出", "Tokens": 30000, "Calls": 80, "Cost": 0.90},
        {"Date": "2024-06-03", "Platform": "Azure", "Project": "代码助手", "Model": "GPT-4o", "Type": "输入", "Tokens": 250000, "Calls": 180, "Amount": 3.75},
        {"Date": "2024-06-03", "Platform": "Azure", "Project": "代码助手", "Model": "GPT-4o", "Type": "输出", "Tokens": 120000, "Calls": 180, "Cost": 4.80},
        {"Date": "2024-06-03", "Platform": "OpenAI", "Project": "文档分析", "Model": "GPT-4-Turbo", "Type": "输入", "Tokens": 100000, "Calls": 90, "Cost": 1.00},
        {"Date": "2024-06-03", "Platform": "OpenAI", "Project": "文档分析", "Model": "GPT-4-Turbo", "Type": "输出", "Tokens": 60000, "Calls": 90, "Cost": 1.80},
    ]
    write_xlsx("test_basic_english.xlsx", headers, records)


# ═══════════════════════════════════════════════════════════════
# 3. 按小时计费 (bill_start != bill_end)
# ═══════════════════════════════════════════════════════════════

def generate_hourly_billing():
    headers = ["开始时间", "截止时间", "平台", "模型", "tokens", "金额"]
    records = [
        {"开始时间": "2024-06-01 00:00:00", "截止时间": "2024-06-01 01:00:00", "平台": "OpenAI", "模型": "GPT-4o", "tokens": 5000, "金额": 0.15},
        {"开始时间": "2024-06-01 01:00:00", "截止时间": "2024-06-01 02:00:00", "平台": "OpenAI", "模型": "GPT-4o", "tokens": 8000, "金额": 0.24},
        {"开始时间": "2024-06-01 02:00:00", "截止时间": "2024-06-01 03:00:00", "平台": "OpenAI", "模型": "GPT-4o", "tokens": 3000, "金额": 0.09},
        {"开始时间": "2024-06-01 08:00:00", "截止时间": "2024-06-01 09:00:00", "平台": "Azure", "模型": "GPT-4-Turbo", "tokens": 12000, "金额": 0.12},
        {"开始时间": "2024-06-01 09:00:00", "截止时间": "2024-06-01 10:00:00", "平台": "Azure", "模型": "GPT-4-Turbo", "tokens": 15000, "金额": 0.15},
    ]
    write_xlsx("test_hourly_billing.xlsx", headers, records)


# ═══════════════════════════════════════════════════════════════
# 4. 含未匹配列 (extra → JSON)
# ═══════════════════════════════════════════════════════════════

def generate_extra_columns():
    headers = ["日期", "平台", "模型", "费用", "备注", "部门", "负责人"]
    records = [
        {"日期": "2024-06-01", "平台": "OpenAI", "模型": "GPT-4o", "费用": 2.00,     "备注": "测试环境", "部门": "AI组", "负责人": "张三"},
        {"日期": "2024-06-02", "平台": "Azure", "模型": "GPT-4-Turbo", "费用": 1.50, "备注": "生产环境", "部门": "后端组", "负责人": "李四"},
        {"日期": "2024-06-03", "平台": "百度文心", "模型": "文心一言4.0", "费用": 3.00, "备注": "POC验证", "部门": "创新组", "负责人": "王五"},
        {"日期": "2024-06-03", "平台": "OpenAI", "模型": "GPT-4o-mini", "费用": 0.80, "备注": "灰度发布", "部门": "AI组", "负责人": "张三"},
    ]
    write_xlsx("test_extra_columns.xlsx", headers, records)


# ═══════════════════════════════════════════════════════════════
# 5. 冲突检测 — 与 test_basic_chinese 部分 key 重叠, 数值不同
# ═══════════════════════════════════════════════════════════════

def generate_conflict():
    """与 basic_chinese 有相同的唯一键但不同的 tokens/cost，用于测试冲突弹窗"""
    headers = ["日期", "平台", "项目", "模型", "类型", "tokens", "调用量", "金额"]
    records = [
        # 完全重叠 — 与 basic_chinese 第1条 key 相同，cost 不同
        {"日期": "2024-06-01", "平台": "OpenAI", "项目": "智能客服", "模型": "GPT-4o", "类型": "输入", "tokens": 155000, "调用量": 120, "金额": 5.50},
        # 完全重叠 — 与 basic_chinese 第5条 key 相同，tokens 不同
        {"日期": "2024-06-02", "平台": "OpenAI", "项目": "智能客服", "模型": "GPT-4o", "类型": "输入", "tokens": 170000, "调用量": 130, "金额": 4.80},
        # 新数据，不冲突
        {"日期": "2024-06-03", "平台": "百度文心", "项目": "智能客服", "模型": "文心一言4.0", "类型": "输入", "tokens": 80000, "调用量": 60, "金额": 1.60},
        {"日期": "2024-06-03", "平台": "百度文心", "项目": "智能客服", "模型": "文心一言4.0", "类型": "输出", "tokens": 45000, "调用量": 60, "金额": 1.80},
    ]
    write_xlsx("test_conflict.xlsx", headers, records)


# ═══════════════════════════════════════════════════════════════
# 6. CSV 文件 — 中文列名 (utf-8 编码)
# ═══════════════════════════════════════════════════════════════

def generate_csv_utf8():
    headers = ["日期", "平台", "项目", "模型", "类型", "tokens", "调用量", "金额"]
    records = [
        {"日期": "2024-06-01", "平台": "OpenAI", "项目": "智能客服", "模型": "GPT-4o", "类型": "输入", "tokens": 120000, "调用量": 100, "金额": 3.60},
        {"日期": "2024-06-01", "平台": "OpenAI", "项目": "智能客服", "模型": "GPT-4o", "类型": "输出", "tokens": 65000, "调用量": 100, "金额": 2.60},
        {"日期": "2024-06-02", "平台": "Azure", "项目": "代码助手", "模型": "GPT-4-Turbo", "类型": "输入", "tokens": 180000, "调用量": 150, "金额": 1.80},
        {"日期": "2024-06-02", "平台": "Azure", "项目": "代码助手", "模型": "GPT-4-Turbo", "类型": "输出", "tokens": 90000, "调用量": 150, "金额": 2.70},
        {"日期": "2024-06-03", "平台": "百度文心", "项目": "文档分析", "模型": "文心一言4.0", "类型": "输入", "tokens": 250000, "调用量": 220, "金额": 5.00},
        {"日期": "2024-06-03", "平台": "百度文心", "项目": "文档分析", "模型": "文心一言4.0", "类型": "输出", "tokens": 140000, "调用量": 220, "金额": 5.60},
    ]
    write_csv("test_csv_utf8.csv", headers, records, encoding="utf-8-sig")


# ═══════════════════════════════════════════════════════════════
# 7. CSV 文件 — GBK 编码 (模拟国内平台导出)
# ═══════════════════════════════════════════════════════════════

def generate_csv_gbk():
    headers = ["日期", "平台", "模型", "费用", "备注"]
    records = [
        {"日期": "2024-06-01", "平台": "OpenAI", "模型": "GPT-4o-mini", "费用": 0.50, "备注": "GBK测试"},
        {"日期": "2024-06-02", "平台": "百度文心", "模型": "文心一言4.0", "费用": 2.00, "备注": "编码探测"},
        {"日期": "2024-06-03", "平台": "Azure", "模型": "GPT-4o", "费用": 3.50, "备注": "兜底测试"},
    ]
    write_csv("test_csv_gbk.csv", headers, records, encoding="gbk")


# ═══════════════════════════════════════════════════════════════
# 8. DeepSeek 官方账单格式 CSV
# ═══════════════════════════════════════════════════════════════

def generate_deepseek_csv():
    headers = ["user_id", "utc_date", "model", "api_key_name", "api_key",
               "type", "price", "amount"]
    records = [
        # user_id 在导入时会被忽略 (归入 extra)
        {"user_id": "u001", "utc_date": "2024-06-01", "model": "deepseek-chat",
         "api_key_name": "default", "api_key": "sk-xxx1",
         "type": "input_cache_hit_tokens", "price": "0.002", "amount": "5000"},
        {"user_id": "u001", "utc_date": "2024-06-01", "model": "deepseek-chat",
         "api_key_name": "default", "api_key": "sk-xxx1",
         "type": "input_cache_miss_tokens", "price": "0.002", "amount": "3000"},
        {"user_id": "u001", "utc_date": "2024-06-01", "model": "deepseek-chat",
         "api_key_name": "default", "api_key": "sk-xxx1",
         "type": "output_tokens", "price": "0.002", "amount": "8000"},
        {"user_id": "u001", "utc_date": "2024-06-01", "model": "deepseek-chat",
         "api_key_name": "default", "api_key": "sk-xxx1",
         "type": "request_count", "price": "0", "amount": "120"},
        {"user_id": "u002", "utc_date": "2024-06-02", "model": "deepseek-reasoner",
         "api_key_name": "prod", "api_key": "sk-xxx2",
         "type": "input_cache_hit_tokens", "price": "0.004", "amount": "10000"},
        {"user_id": "u002", "utc_date": "2024-06-02", "model": "deepseek-reasoner",
         "api_key_name": "prod", "api_key": "sk-xxx2",
         "type": "output_tokens", "price": "0.004", "amount": "6000"},
        {"user_id": "u002", "utc_date": "2024-06-02", "model": "deepseek-reasoner",
         "api_key_name": "prod", "api_key": "sk-xxx2",
         "type": "request_count", "price": "0", "amount": "50"},
    ]
    write_csv("test_deepseek.csv", headers, records, encoding="utf-8-sig")


# ═══════════════════════════════════════════════════════════════
# 9. 新表头格式: 资源名称 模型 类型 token 账单开始时间 账单结束时间 费用(元)
# ═══════════════════════════════════════════════════════════════

def generate_new_header_format():
    headers = ["资源名称", "模型", "类型", "token", "账单开始时间", "账单结束时间", "费用(元)"]
    records = [
        {"资源名称": "智障助手v2", "模型": "GPT-4o", "类型": "输入", "token": 150000,
         "账单开始时间": "2024-06-01", "账单结束时间": "2024-06-01", "费用(元)": 4.50},
        {"资源名称": "智障助手v2", "模型": "GPT-4o", "类型": "输出", "token": 80000,
         "账单开始时间": "2024-06-01", "账单结束时间": "2024-06-01", "费用(元)": 3.20},
        {"资源名称": "代码助手", "模型": "GPT-4-Turbo", "类型": "输入", "token": 200000,
         "账单开始时间": "2024-06-01", "账单结束时间": "2024-06-01", "费用(元)": 2.00},
        {"资源名称": "代码助手", "模型": "GPT-4-Turbo", "类型": "输出", "token": 95000,
         "账单开始时间": "2024-06-01", "账单结束时间": "2024-06-01", "费用(元)": 2.85},
        {"资源名称": "文档分析", "模型": "DeepSeek-V3", "类型": "输入(缓存命中)", "token": 50000,
         "账单开始时间": "2024-06-02", "账单结束时间": "2024-06-02", "费用(元)": 0.10},
    ]
    write_csv("test_new_header.csv", headers, records, encoding="utf-8-sig")


# ═══════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════

def main():
    print("生成测试数据...\n")
    generate_basic_chinese()
    generate_basic_english()
    generate_hourly_billing()
    generate_extra_columns()
    generate_conflict()
    generate_csv_utf8()
    generate_csv_gbk()
    generate_deepseek_csv()
    generate_new_header_format()
    print(f"\n完成！所有文件已输出到: {TEST_DIR}")


if __name__ == "__main__":
    main()
